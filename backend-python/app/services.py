from __future__ import annotations

import json
import smtplib
from datetime import datetime
from email.mime.text import MIMEText
from pathlib import Path

import redis
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.config import get_settings
from app.llm import LlmClient
from app.models import (
    ChatMessage,
    ChatSession,
    EmailAlertLog,
    ExcelExportLog,
    KnowledgeChunk,
    KnowledgeDocument,
    RiskRecord,
    SysUser,
    UserMemory,
    WorkflowRecord,
)
from app.prompts import INTENT_SYSTEM, intent_user, reply_system, reply_user
from app.retrieval import HybridRetrievalPipeline, QdrantVectorStore, chunk_text, stable_memory_key
from app.schemas import ClassificationResult, RagReference
from app.utils import ensure_parent, model_to_dict, safe_json

RISK_ORDER = {"LOW": 0, "MEDIUM": 1, "HIGH": 2}
HIGH_RISK_KEYWORDS = [
    "想死",
    "不想活",
    "活不下去",
    "自杀",
    "结束生命",
    "伤害自己",
    "割腕",
    "跳楼",
    "没人救我",
    "我撑不住了",
    "撑不住了",
    "报复",
    "杀人",
    "伤害别人",
]


class RiskRuleService:
    def detect(self, message: str) -> tuple[str, str, str | None]:
        normalized = (message or "").replace(" ", "")
        for keyword in HIGH_RISK_KEYWORDS:
            if keyword in normalized:
                risk_type = "violence" if any(word in keyword for word in ["杀", "报复", "别人"]) else "self_harm"
                return "HIGH", risk_type, keyword
        if any(word in normalized for word in ["崩溃", "绝望", "睡不着"]):
            return "MEDIUM", "distress", None
        return "LOW", "none", None


class MemoryService:
    def __init__(self) -> None:
        self.settings = get_settings()
        self.redis_client = redis.Redis.from_url(self.settings.redis_url, decode_responses=True)

    def recent_context(self, user_id: int, session_id: int) -> str:
        key = self._key(user_id, session_id)
        try:
            turns = self.redis_client.lrange(key, 0, 9)
        except Exception:
            return "暂无短期记忆。"
        if not turns:
            return "暂无短期记忆。"
        return "\n".join(reversed(turns))

    def append_turn(self, user_id: int, session_id: int, user_message: str, ai_reply: str) -> None:
        key = self._key(user_id, session_id)
        try:
            self.redis_client.lpush(key, f"用户：{user_message}\n助手：{ai_reply}")
            self.redis_client.ltrim(key, 0, 9)
            self.redis_client.expire(key, 7 * 24 * 3600)
        except Exception:
            return

    def long_context(self, db: Session, user_id: int) -> str:
        memories = (
            db.query(UserMemory)
            .filter(UserMemory.user_id == user_id)
            .order_by(UserMemory.update_time.desc())
            .limit(8)
            .all()
        )
        if not memories:
            return "暂无长期记忆。"
        return "\n".join(f"- {memory.content}" for memory in memories)

    def save_long_memory(
        self,
        db: Session,
        user_id: int,
        session_id: int,
        user_message: str,
        intent: str,
        risk_level: str,
    ) -> None:
        if intent == "CHAT" or risk_level == "HIGH":
            return
        content = user_message.strip()
        if len(content) < 8:
            return
        key = stable_memory_key(content)
        existing = db.query(UserMemory).filter(UserMemory.user_id == user_id, UserMemory.memory_key == key).first()
        if existing:
            existing.last_used_time = datetime.now()
            return
        db.add(
            UserMemory(
                user_id=user_id,
                memory_key=key,
                memory_type="conversation_preference",
                content=content[:300],
                source_session_id=session_id,
                importance=3,
                last_used_time=datetime.now(),
            )
        )

    def _key(self, user_id: int, session_id: int) -> str:
        return f"mindcare:memory:short:{user_id}:{session_id}"


class KnowledgeService:
    def __init__(self, vector_store: QdrantVectorStore) -> None:
        self.vector_store = vector_store

    def upload(self, db: Session, file_name: str, content: str) -> KnowledgeDocument:
        chunks = chunk_text(content)
        document = KnowledgeDocument(file_name=file_name, content=content, chunk_count=len(chunks))
        db.add(document)
        db.flush()
        chunk_models = []
        for index, chunk in enumerate(chunks):
            model = KnowledgeChunk(
                document_id=document.id,
                document_name=file_name,
                chunk_index=index,
                content=chunk,
            )
            db.add(model)
            chunk_models.append(model)
        db.flush()
        self.vector_store.upsert_chunks(chunk_models)
        db.commit()
        db.refresh(document)
        return document


class ExcelService:
    HEADERS = [
        "记录ID",
        "用户ID",
        "会话ID",
        "用户问题",
        "意图类型",
        "风险类型",
        "风险等级",
        "是否命中RAG",
        "RAG参考片段",
        "AI回复",
        "是否发送邮件",
        "创建时间",
    ]

    def __init__(self) -> None:
        self.settings = get_settings()

    def append_workflow_record(self, db: Session, record: WorkflowRecord) -> Path:
        path = ensure_parent(self.settings.excel_workflow_path)
        if path.exists():
            workbook = load_workbook(path)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "workflow"
            sheet.append(self.HEADERS)
        sheet.append(self._row(record))
        workbook.save(path)
        record.excel_exported = True
        db.add(ExcelExportLog(record_id=record.id, file_path=str(path), export_status="SUCCESS"))
        return path

    def export_all(self, db: Session, records: list[WorkflowRecord]) -> Path:
        file_name = f"workflow-records-{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        path = ensure_parent(Path(self.settings.excel_export_dir) / file_name)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "workflow"
        sheet.append(self.HEADERS)
        for record in records:
            sheet.append(self._row(record))
        workbook.save(path)
        db.add(ExcelExportLog(record_id=None, file_path=str(path), export_status="SUCCESS"))
        db.commit()
        return path

    def _row(self, record: WorkflowRecord) -> list:
        return [
            record.id,
            record.user_id,
            record.session_id,
            record.user_message,
            record.intent,
            record.risk_type,
            record.risk_level,
            "是" if record.rag_hit else "否",
            record.rag_references,
            record.ai_reply,
            "是" if record.email_sent else "否",
            record.create_time.isoformat(sep=" ") if record.create_time else "",
        ]


class EmailService:
    def __init__(self) -> None:
        self.settings = get_settings()

    def send_risk_alert(self, db: Session, risk_record: RiskRecord) -> bool:
        student = db.get(SysUser, risk_record.user_id)
        student_name = self._first(student.real_name if student else None, student.username if student else None, "未知学生")
        receiver = self._resolve_receiver(db)
        subject = f"心理陪伴助手高风险预警 - {student_name}"
        content = f"""检测到高风险会话，请负责老师尽快关注。

【学生信息】
学生姓名：{student_name}
学院：{self._first(student.college if student else None, "未配置学院")}
学生账号：{self._first(student.username if student else None, "未配置账号")}
学生邮箱：{self._first(student.email if student else None, "未配置邮箱")}
学生ID：{risk_record.user_id}

【负责老师】
老师姓名：{self.settings.mail_teacher_name}
所属部门：{self.settings.mail_teacher_department}
预警接收邮箱：{receiver}

【风险信息】
会话ID：{risk_record.session_id}
风险等级：{risk_record.risk_level}
风险类型：{risk_record.risk_type or "未分类"}

【学生原文】
{risk_record.user_message}

【AI安全回复】
{risk_record.ai_reply}
"""
        return self._send(db, risk_record.id, receiver, subject, content)

    def send_test(self, db: Session, receiver: str | None, subject: str | None, content: str | None) -> bool:
        return self._send(
            db,
            None,
            receiver or self._resolve_receiver(db),
            subject or "心理陪伴助手邮件测试",
            content or "这是一封邮件预警通道测试邮件。",
        )

    def _send(self, db: Session, risk_record_id: int | None, receiver: str, subject: str, content: str) -> bool:
        if not self.settings.mail_enabled:
            self._log(db, risk_record_id, receiver, subject, content, "SKIPPED", "MAIL_ENABLED=false")
            return False
        try:
            message = MIMEText(content, "plain", "utf-8")
            message["From"] = self.settings.mail_from or self.settings.mail_username
            message["To"] = receiver
            message["Subject"] = subject
            if self.settings.mail_use_ssl:
                server = smtplib.SMTP_SSL(self.settings.mail_host, self.settings.mail_port, timeout=20)
            else:
                server = smtplib.SMTP(self.settings.mail_host, self.settings.mail_port, timeout=20)
            with server:
                if self.settings.mail_use_tls:
                    server.starttls()
                server.login(self.settings.mail_username, self.settings.mail_password)
                server.sendmail(message["From"], [receiver], message.as_string())
            self._log(db, risk_record_id, receiver, subject, content, "SUCCESS", None)
            return True
        except Exception as exc:
            self._log(db, risk_record_id, receiver, subject, content, "FAILED", str(exc))
            return False

    def _log(
        self,
        db: Session,
        risk_record_id: int | None,
        receiver: str,
        subject: str,
        content: str,
        status: str,
        error: str | None,
    ) -> None:
        db.add(
            EmailAlertLog(
                risk_record_id=risk_record_id,
                receiver=receiver,
                subject=subject,
                content=content,
                send_status=status,
                error_message=error,
            )
        )

    def _resolve_receiver(self, db: Session) -> str:
        if self.settings.mail_alert_receiver and not self.settings.mail_alert_receiver.endswith("@example.com"):
            return self.settings.mail_alert_receiver
        admin = db.query(SysUser).filter(SysUser.role == "ADMIN").first()
        return admin.email if admin and admin.email else self.settings.mail_alert_receiver

    def _first(self, *values: str | None) -> str:
        for value in values:
            if value and value.strip():
                return value.strip()
        return ""


class ChatWorkflowService:
    def __init__(self, retriever: HybridRetrievalPipeline, llm_client: LlmClient) -> None:
        self.settings = get_settings()
        self.retriever = retriever
        self.llm_client = llm_client
        self.risk_rules = RiskRuleService()
        self.memory = MemoryService()
        self.excel = ExcelService()
        self.email = EmailService()

    def process_message(self, db: Session, user_id: int, session_id: int | None, message: str) -> dict:
        session = self._ensure_session(db, user_id, session_id, message)
        self._save_message(db, session.id, "USER", message, None, None)
        actions = ["RAG_DENSE_QDRANT", "RAG_KEYWORD_BM25", "RAG_RRF_FUSION", "RAG_CROSS_ENCODER_RERANK"]
        references = self.retriever.search(db, message)
        rag_context = self._rag_context(references)

        short_memory = self.memory.recent_context(user_id, session.id)
        long_memory = self.memory.long_context(db, user_id)
        classification = self._classify(message, rag_context)
        rule_level, rule_type, _ = self.risk_rules.detect(message)
        final_risk = self._max_risk(classification.riskLevel, rule_level)
        final_intent = classification.intent
        risk_type = rule_type if rule_level != "LOW" else classification.riskType or "none"
        if final_intent == "HIGH_RISK" or final_risk == "HIGH":
            final_intent = "HIGH_RISK"
            final_risk = "HIGH"
            actions.append("RULE_OR_LLM_HIGH_RISK_CHECK")

        reply = self._reply(final_intent, message, rag_context, short_memory, long_memory)
        self._save_message(db, session.id, "ASSISTANT", reply, final_intent, final_risk)
        self.memory.append_turn(user_id, session.id, message, reply)
        self.memory.save_long_memory(db, user_id, session.id, message, final_intent, final_risk)

        response = {
            "sessionId": session.id,
            "reply": reply,
            "intent": final_intent,
            "riskLevel": final_risk,
            "riskType": risk_type,
            "ragHit": bool(references),
            "references": [reference.model_dump() for reference in references],
            "actions": actions,
        }
        if final_intent == "CHAT":
            db.commit()
            return response

        workflow_record = WorkflowRecord(
            user_id=user_id,
            session_id=session.id,
            user_message=message,
            intent=final_intent,
            risk_type=risk_type,
            risk_level=final_risk,
            rag_hit=bool(references),
            rag_references=safe_json([reference.model_dump() for reference in references]),
            ai_reply=reply,
            excel_exported=False,
            email_sent=False,
        )
        db.add(workflow_record)
        db.flush()
        actions.append("SAVE_WORKFLOW_RECORD")

        if final_intent == "HIGH_RISK":
            risk_record = RiskRecord(
                user_id=user_id,
                session_id=session.id,
                user_message=message,
                risk_type=risk_type,
                risk_level=final_risk,
                ai_reply=reply,
                handled=False,
            )
            db.add(risk_record)
            db.flush()
            actions.append("SAVE_RISK_RECORD")
            self.excel.append_workflow_record(db, workflow_record)
            actions.append("WRITE_EXCEL")
            workflow_record.email_sent = self.email.send_risk_alert(db, risk_record)
            actions.append("SEND_EMAIL_ALERT")
        else:
            self.excel.append_workflow_record(db, workflow_record)
            actions.append("WRITE_EXCEL")
        db.commit()
        response["actions"] = actions
        return response

    def _ensure_session(self, db: Session, user_id: int, session_id: int | None, message: str) -> ChatSession:
        if session_id:
            existing = db.get(ChatSession, session_id)
            if existing and existing.user_id == user_id:
                return existing
        session = ChatSession(user_id=user_id, title=message[:18] or "新会话")
        db.add(session)
        db.flush()
        return session

    def _save_message(self, db: Session, session_id: int, role: str, content: str, intent: str | None, risk: str | None) -> None:
        db.add(ChatMessage(session_id=session_id, role=role, content=content, intent=intent, risk_level=risk))

    def _classify(self, message: str, rag_context: str) -> ClassificationResult:
        try:
            data = self.llm_client.chat_json(INTENT_SYSTEM, intent_user(message, rag_context))
            return ClassificationResult(**data)
        except Exception:
            rule_level, rule_type, _ = self.risk_rules.detect(message)
            if rule_level == "HIGH":
                return ClassificationResult(intent="HIGH_RISK", riskLevel="HIGH", riskType=rule_type, reason="规则命中高风险关键词")
            if any(word in message for word in ["什么", "如何", "方法", "概念", "为什么", "哪些"]):
                return ClassificationResult(intent="KNOWLEDGE", riskLevel=rule_level, riskType=rule_type, reason="启发式识别为知识问答")
            if any(word in message for word in ["压力", "焦虑", "难受", "失眠", "关系", "情绪", "累"]):
                return ClassificationResult(intent="CONSULT", riskLevel=rule_level, riskType=rule_type, reason="启发式识别为咨询")
            return ClassificationResult(intent="CHAT", riskLevel=rule_level, riskType=rule_type, reason="启发式识别为闲聊")

    def _reply(self, intent: str, message: str, rag_context: str, short_memory: str, long_memory: str) -> str:
        try:
            return self.llm_client.chat(reply_system(intent), reply_user(message, rag_context, short_memory, long_memory))
        except Exception:
            return self._fallback_reply(intent)

    def _fallback_reply(self, intent: str) -> str:
        if intent == "HIGH_RISK":
            return "我听见你现在很痛苦。请立刻联系身边可信任的人陪着你，并尽快联系当地紧急服务或专业心理危机热线。先让自己远离可能造成伤害的物品，不要一个人扛着。"
        if intent == "CONSULT":
            return "听起来你最近承受了不少压力。可以先把今晚必须做的事减到最少，做几轮缓慢呼吸，把担心的事情写下来留到明天处理。如果这种状态持续，也建议联系专业支持。"
        if intent == "KNOWLEDGE":
            return "当前知识库中没有找到明确依据。一般来说，规律作息、放松训练、适度运动和记录情绪触发点，常被用于缓解压力与焦虑体验。"
        return "我在呢。无聊的时候可以先挑一件很小的事做做，比如听首歌、倒杯水，或者和我随便聊两句。"

    def _rag_context(self, references: list[RagReference]) -> str:
        if not references:
            return "当前知识库没有找到明确依据。"
        return "\n\n".join(f"[{idx}] 文档：{ref.documentName}\n{ref.content}" for idx, ref in enumerate(references, start=1))

    def _max_risk(self, left: str, right: str) -> str:
        return left if RISK_ORDER.get(left, 0) >= RISK_ORDER.get(right, 0) else right


def seed_default_users(db: Session) -> None:
    if not db.query(SysUser).filter(SysUser.username == "admin").first():
        db.add(
            SysUser(
                username="admin",
                password="{noop}admin123",
                real_name="王老师",
                department="学生心理健康中心",
                email="admin@example.com",
                role="ADMIN",
            )
        )
    if not db.query(SysUser).filter(SysUser.username == "user").first():
        db.add(
            SysUser(
                username="user",
                password="{noop}user123",
                real_name="张同学",
                college="人工智能学院",
                email="student@example.com",
                role="USER",
            )
        )
    db.commit()


def to_api_list(items: list) -> list[dict]:
    return [model_to_dict(item) for item in items]
